from openpyxl import load_workbook

def get_test_data(file_path: str, sheet_name: str) -> list[dict]:
    """
    Reads an Excel sheet and returns a list of dictionaries (one per row).
    If a cell contains comma-separated values, it converts it to a list of strings.
    """
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    headers = [cell.value for cell in sheet[1]]
    all_rows = []

    for row in range(2, sheet.max_row + 1):
        row_data = {}
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row, column=col).value

            # If value is string and contains ',', convert to list
            if isinstance(value, str) and ',' in value:
                row_data[headers[col - 1]] = [v.strip() for v in value.split(',')]
            else:
                row_data[headers[col - 1]] = value
        all_rows.append(row_data)
    return all_rows
